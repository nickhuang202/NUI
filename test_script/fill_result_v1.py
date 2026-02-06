#!/usr/bin/env python3
"""
Script to fill EVT template with test results from hwtest CSV
Supports CSV template input and Excel output with multiple sheets
"""

import csv
import sys
import os
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# ======================
# Argument Validation
# ======================
if len(sys.argv) != 6:
    print(
        "Usage:\n"
        "  python fill_evt_result.py <hwtest.csv> <evt_template.xlsx> <output.xlsx> <column_index> <commit_id>\n\n"
        "Column index:\n"
        "  1 = Optics Config 1 (optics_one)\n"
        "  2 = Optics Config 2 (optics_two)\n"
        "  3 = DAC Config 1 (copper)\n"
        "  4 = AEC Config 1 (copper)\n"
        "  5 = Accton Config 1 (800G)\n"
        "  6 = Accton Config 2 (400G)\n\n"
        "Commit ID:\n"
        "  Used as sheet name. If sheet exists, update it; otherwise create new sheet.\n\n"
        "Example:\n"
        "  python fill_evt_result.py hwtest.csv template.xlsx output.xlsx 5 abc123def\n\n"
        "Note:\n"
        "  - Template file should be Excel (.xlsx) format\n"
        "  - Output file will be Excel (.xlsx) format\n"
    )
    sys.exit(1)

HWTEST_FILE = sys.argv[1]
EVT_FILE = sys.argv[2]
OUTPUT_FILE = sys.argv[3]
COLUMN_INDEX = sys.argv[4]
COMMIT_ID = sys.argv[5]

# ======================
# Column Configuration
# ======================
HWTEST_NAME_COL = "Test Name"
HWTEST_RESULT_COL = "Result"
EVT_NAME_COL = "Test Name"

COLUMN_MAP = {
    "1": "Optics Config 1 (optics_one)",
    "2": "Optics Config 2 (optics_two)",
    "3": "DAC Config 1 (copper)",
    "4": "AEC Config 1 (copper)",
    "5": "Accton Config 1 (800G)",
    "6": "Accton Config 2 (400G)",
}

if COLUMN_INDEX not in COLUMN_MAP:
    print("ERROR: Column index must be 1, 2, 3, 4, 5, or 6")
    sys.exit(1)

TARGET_COLUMN = COLUMN_MAP[COLUMN_INDEX]

# Validate sheet name (Excel has restrictions)
if len(COMMIT_ID) > 31:
    print(f"ERROR: Commit ID too long (max 31 characters): {COMMIT_ID}")
    sys.exit(1)

invalid_chars = ['/', '\\', '?', '*', '[', ']', ':']
if any(char in COMMIT_ID for char in invalid_chars):
    print(f"ERROR: Commit ID contains invalid characters: {COMMIT_ID}")
    print(f"Invalid characters: {', '.join(invalid_chars)}")
    sys.exit(1)

# ======================
# Helper Functions
# ======================
def normalize_test_name(name: str) -> str:
    """
    Remove boot prefix from test name
    Examples:
      warm_boot.test_xyz -> test_xyz
      cold_boot.test_xyz -> test_xyz
      reboot.test_xyz -> test_xyz
    """
    for prefix in ("warm_boot.", "cold_boot.", "reboot."):
        if name.startswith(prefix):
            return name[len(prefix):]
    return name

def normalize_result(result: str) -> str:
    """
    Normalize result to PASS or FAIL
    PASS/OK -> PASS
    Everything else -> FAIL
    """
    result = result.strip().upper()
    if result in ("PASS", "OK"):
        return "PASS"
    return "FAIL"

def load_csv_template(csv_file):
    """Load CSV template and return headers and rows"""
    with open(csv_file, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames
        rows = list(reader)
    return headers, rows

def create_sheet_from_csv(wb, sheet_name, headers, rows):
    """Create a new sheet in workbook from CSV data"""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Write rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
    
    return ws

def copy_sheet_structure(wb, source_sheet, new_sheet_name):
    """Copy structure from source sheet to new sheet"""
    if new_sheet_name in wb.sheetnames:
        return wb[new_sheet_name]
    
    ws = wb.create_sheet(title=new_sheet_name)
    
    # Copy all cells
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
            # Copy cell formatting if needed
            if cell.has_style:
                new_cell.font = cell.font.copy()
                new_cell.border = cell.border.copy()
                new_cell.fill = cell.fill.copy()
                new_cell.number_format = cell.number_format
                new_cell.protection = cell.protection.copy()
                new_cell.alignment = cell.alignment.copy()
    
    return ws

def find_column_index(sheet, column_name):
    """Find the column index (1-based) for a given column name"""
    for col_idx, cell in enumerate(sheet[1], start=1):
        if cell.value == column_name:
            return col_idx
    return None

# ======================
# Read hwtest CSV
# ======================
print(f"Reading hwtest file: {HWTEST_FILE}")
hw_map = {}

try:
    with open(HWTEST_FILE, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        
        if HWTEST_NAME_COL not in reader.fieldnames:
            print(f"ERROR: Column '{HWTEST_NAME_COL}' not found in hwtest CSV")
            sys.exit(1)
        
        if HWTEST_RESULT_COL not in reader.fieldnames:
            print(f"ERROR: Column '{HWTEST_RESULT_COL}' not found in hwtest CSV")
            sys.exit(1)
        
        for row in reader:
            raw_name = row.get(HWTEST_NAME_COL, "").strip()
            raw_result = row.get(HWTEST_RESULT_COL, "").strip()
            
            if not raw_name:
                continue
            
            name = normalize_test_name(raw_name)
            result = normalize_result(raw_result)
            
            # FAIL takes priority
            if name in hw_map and hw_map[name] == "FAIL":
                continue
            
            hw_map[name] = result

    print(f"  Loaded {len(hw_map)} test results from hwtest file")
    
except FileNotFoundError:
    print(f"ERROR: hwtest file not found: {HWTEST_FILE}")
    sys.exit(1)
except Exception as e:
    print(f"ERROR reading hwtest file: {e}")
    sys.exit(1)

# ======================
# Load Excel Template
# ======================
print(f"\nReading EVT template: {EVT_FILE}")

try:
    # Load Excel template directly
    template_wb = load_workbook(EVT_FILE)
    template_ws = template_wb.active
    
    # Extract headers from first row
    template_headers = [cell.value for cell in template_ws[1]]
    
    if EVT_NAME_COL not in template_headers:
        print(f"ERROR: Column '{EVT_NAME_COL}' not found in template")
        sys.exit(1)
    
    if TARGET_COLUMN not in template_headers:
        print(f"ERROR: Column '{TARGET_COLUMN}' not found in template")
        print(f"Available columns: {', '.join(str(h) for h in template_headers if h)}")
        sys.exit(1)
    
    print(f"  Loaded template with {template_ws.max_row - 1} data rows")

except FileNotFoundError:
    print(f"ERROR: Template file not found: {EVT_FILE}")
    sys.exit(1)
except Exception as e:
    print(f"ERROR reading template file: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# ======================
# Load or Create Excel Workbook
# ======================
print(f"\nProcessing Excel output: {OUTPUT_FILE}")

try:
    if os.path.exists(OUTPUT_FILE):
        print(f"  Loading existing workbook: {OUTPUT_FILE}")
        wb = load_workbook(OUTPUT_FILE)
        
        # Check if sheet with commit ID exists
        if COMMIT_ID in wb.sheetnames:
            print(f"  Sheet '{COMMIT_ID}' exists - will update it")
            ws = wb[COMMIT_ID]
        else:
            print(f"  Sheet '{COMMIT_ID}' not found - copying from template")
            # Copy the template sheet
            ws = wb.copy_worksheet(template_ws)
            ws.title = COMMIT_ID
    else:
        print(f"  Creating new workbook from template: {OUTPUT_FILE}")
        # Use template workbook as base
        wb = load_workbook(EVT_FILE)
        # Rename first sheet to commit ID
        ws = wb.active
        ws.title = COMMIT_ID
        print(f"  Created new sheet '{COMMIT_ID}' from template")

except Exception as e:
    print(f"ERROR processing Excel file: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# ======================
# Find Required Columns
# ======================
print(f"\nLocating columns in sheet '{COMMIT_ID}'...")

test_name_col_idx = find_column_index(ws, EVT_NAME_COL)
if not test_name_col_idx:
    print(f"ERROR: Column '{EVT_NAME_COL}' not found in sheet")
    sys.exit(1)
print(f"  Test Name column: {get_column_letter(test_name_col_idx)} (index {test_name_col_idx})")

target_col_idx = find_column_index(ws, TARGET_COLUMN)
if not target_col_idx:
    print(f"ERROR: Column '{TARGET_COLUMN}' not found in sheet")
    available_cols = [cell.value for cell in ws[1] if cell.value]
    print(f"Available columns: {', '.join(available_cols)}")
    sys.exit(1)
print(f"  Target column '{TARGET_COLUMN}': {get_column_letter(target_col_idx)} (index {target_col_idx})")

# ======================
# Fill PASS/FAIL Results
# ======================
print(f"\nFilling results into column: {TARGET_COLUMN}")

matched_count = 0
unmatched_count = 0
total_rows = ws.max_row

for row_idx in range(2, total_rows + 1):  # Start from row 2 (skip header)
    test_name_cell = ws.cell(row=row_idx, column=test_name_col_idx)
    test_name = str(test_name_cell.value).strip() if test_name_cell.value else ""
    
    if not test_name or test_name == "None":
        continue
    
    if test_name in hw_map:
        result = hw_map[test_name]
        target_cell = ws.cell(row=row_idx, column=target_col_idx, value=result)
        
        # Apply red color to FAIL results
        if result == "FAIL":
            target_cell.font = Font(color="FF0000", bold=True)
        
        matched_count += 1
    else:
        unmatched_count += 1

print(f"  Matched tests: {matched_count}")
if unmatched_count > 0:
    print(f"  Unmatched tests: {unmatched_count} (no result found in hwtest file)")

# ======================
# Save Output File
# ======================
print(f"\nSaving output file: {OUTPUT_FILE}")

try:
    wb.save(OUTPUT_FILE)
    
    print("\n" + "="*60)
    print("SUCCESS: EVT result filled successfully")
    print("="*60)
    print(f"Sheet name    : {COMMIT_ID}")
    print(f"Filled column : {TARGET_COLUMN}")
    print(f"Output file   : {OUTPUT_FILE}")
    print(f"Tests matched : {matched_count}/{total_rows - 1}")
    print(f"Total sheets  : {len(wb.sheetnames)}")
    print("="*60)

except Exception as e:
    print(f"ERROR saving output file: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

